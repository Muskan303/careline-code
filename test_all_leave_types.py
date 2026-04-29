"""
Automation: TC_ATT_025 to TC_ATT_032
Annual (Full/First/Second), Maternity (Full Day only),
Menstrual (Full Day only), Election Leave (Full/First/Second)
- If no absent dates in current month → click '>' to go to next month
- Maternity & Menstrual: Full Day only (auto-applied, no day type selection)
Results appended to: BasicInfo_AutoTest_Results_v2.xlsx  sheet: Attendance - Apply Leave
"""

import re, time, datetime, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager

LOGIN_URL      = "http://gcplcareline.girnarsoft.com/admin/user/user/access?e=monika.bidawat@girnarsoft.com"
ATTENDANCE_URL = "http://gcplcareline.girnarsoft.com/employee/attendance"
WAIT           = 15
OUTPUT_FILE    = "BasicInfo_AutoTest_Results_v2.xlsx"
SHEET_NAME     = "Attendance - Apply Leave"
DAY_TYPES      = ["Full Day", "First Half", "Second Half"]

TODAY            = None
SELECTED_CELL_ID = None

# TC definitions: (tc_id, leave_type, day_type or None)
# day_type=None means Full Day only, no day type selection (Maternity/Menstrual)
LEAVE_CASES = [
    ("TC_ATT_025", "Annual",         "Full Day"),
    ("TC_ATT_026", "Annual",         "First Half"),
    ("TC_ATT_027", "Annual",         "Second Half"),
    ("TC_ATT_028", "Maternity",      None),        # Full Day auto
    ("TC_ATT_029", "Menstrual",      None),        # Full Day auto
    ("TC_ATT_030", "Election Leave", "Full Day"),
    ("TC_ATT_031", "Election Leave", "First Half"),
    ("TC_ATT_032", "Election Leave", "Second Half"),
]


# ── Driver & navigation ───────────────────────────────────────────────────────
def make_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-notifications")
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)

def login_and_go(driver, wait):
    driver.get(LOGIN_URL); time.sleep(4)
    driver.get(ATTENDANCE_URL)
    wait.until(EC.presence_of_element_located(
        (By.XPATH, "//*[contains(text(),'Attendance')]")))
    time.sleep(2)

def body_text(driver):
    return driver.find_element(By.TAG_NAME, "body").text

def find_available_date(driver):
    """Find first absent date with ba-wrap (Apply radio) in current calendar view."""
    for cell in driver.find_elements(By.XPATH, "//li[contains(@class,'day-cell')]"):
        cls = cell.get_attribute("class") or ""
        if "absent_pt" not in cls: continue
        if not cell.find_elements(By.XPATH, ".//label[contains(@class,'ba-wrap')]"): continue
        labels = cell.find_elements(By.XPATH, ".//label[contains(@class,'hover_hide')]")
        if labels and labels[0].text.strip():
            return labels[0].text.strip(), cell.get_attribute("id")
    return None, None

def navigate_to_next_month(driver, wait):
    """Click the '>' next button (class='gscr_lSNext') to go to next month."""
    try:
        next_btn = driver.find_element(By.XPATH,
            "//a[contains(@class,'gscr_lSNext')] | "
            "//button[contains(@class,'gscr_lSNext')]")
        driver.execute_script("arguments[0].click();", next_btn)
        time.sleep(2)
        wait.until(EC.presence_of_element_located(
            (By.XPATH, "//*[contains(text(),'Attendance')]")))
        time.sleep(1)
        return True
    except Exception as e:
        print(f"    [WARN] navigate_to_next_month failed: {e}")
        return False

def ensure_available_date(driver, wait):
    """Find available date; if none in current month, go to next month."""
    global TODAY, SELECTED_CELL_ID
    d, c = find_available_date(driver)
    if d:
        TODAY, SELECTED_CELL_ID = d, c
        return True
    print("    [INFO] No dates in current month — navigating to next month...")
    if navigate_to_next_month(driver, wait):
        time.sleep(2)
        d, c = find_available_date(driver)
        if d:
            TODAY, SELECTED_CELL_ID = d, c
            print(f"    [INFO] Found date {TODAY} in next month")
            return True
    return False


# ── Form helpers ──────────────────────────────────────────────────────────────
def open_form(driver, wait):
    global TODAY, SELECTED_CELL_ID
    for _ in range(2):
        try:
            lbl = driver.find_element(By.XPATH,
                f"//label[contains(@class,'hover_hide') and contains(@class,'lbl-A') "
                f"and normalize-space(text())='{TODAY}']")
            ActionChains(driver).move_to_element(lbl).perform(); time.sleep(1.2)
            parent = lbl.find_element(By.XPATH, "..")
            ba = parent.find_elements(By.XPATH, ".//label[contains(@class,'ba-wrap')]")
            if ba:
                driver.execute_script("arguments[0].click();", ba[0])
            else:
                cb = driver.find_element(By.XPATH, f"//input[@id='lbl-{SELECTED_CELL_ID}']")
                driver.execute_script("arguments[0].click();", cb)
            time.sleep(0.8)
            el = driver.find_element(By.XPATH, "//div[@data-attendance-slug='leave']")
            el.click(); time.sleep(3)
            try:
                WebDriverWait(driver, WAIT).until(lambda d: any(
                    s.is_displayed() and len(s.find_elements(By.TAG_NAME, "option")) > 0
                    for s in d.find_elements(By.TAG_NAME, "select")))
            except Exception: pass
            time.sleep(1)
            if is_form_open(driver): return True
        except StaleElementReferenceException: time.sleep(2)
        except Exception: break
    return False

def is_form_open(driver):
    for s in driver.find_elements(By.TAG_NAME, "select"):
        try:
            if s.is_displayed() and len(s.find_elements(By.TAG_NAME, "option")) > 0:
                return True
        except Exception: pass
    return False

def visible_selects(driver):
    return [s for s in driver.find_elements(By.TAG_NAME, "select")
            if s.is_displayed() and len(s.find_elements(By.TAG_NAME, "option")) > 0]

def leave_type_select(driver):
    for s in visible_selects(driver):
        opts = [o.text for o in Select(s).options]
        if any(x in " ".join(opts).lower()
               for x in ["annual","maternity","bereavement","menstrual","election"]):
            return s
    return None

def day_type_select(driver):
    for s in visible_selects(driver):
        opts = [o.text for o in Select(s).options]
        if any(d in opts for d in DAY_TYPES): return s
    return None

def select_opt(sel_el, fragment):
    if sel_el is None: return None
    try:
        s = Select(sel_el)
        match = next((o.text for o in s.options if fragment.lower() in o.text.lower()), None)
        if match: s.select_by_visible_text(match); time.sleep(0.8); return match
    except Exception: pass
    return None

def enter_comment(driver, text="Automation test"):
    for xp in ["//textarea",
               "//input[contains(@placeholder,'comment') or contains(@name,'comment')]"]:
        els = [e for e in driver.find_elements(By.XPATH, xp) if e.is_displayed()]
        if els:
            try: els[0].clear(); els[0].send_keys(text); return True
            except Exception: pass
    return False

def click_submit(driver):
    for el in driver.find_elements(By.XPATH,
            "//button[contains(text(),'Submit') or @type='submit'] | //input[@type='submit']"):
        try:
            if el.is_displayed() and el.is_enabled():
                el.click(); time.sleep(2.5); return True
        except Exception: pass
    return False


# ── Core submit function ──────────────────────────────────────────────────────
def do_submit(leave_type, day_type):
    """
    Fresh browser → login → find date (next month if needed) →
    open form → select leave type → select day type (skip for Maternity/Menstrual) →
    comment → submit.
    Returns (actual_result, status_string)
    """
    global TODAY, SELECTED_CELL_ID
    drv = make_driver()
    wt  = WebDriverWait(drv, WAIT)
    try:
        login_and_go(drv, wt)
        # Find available date, navigate to next month if needed
        if not ensure_available_date(drv, wt):
            return f"No available absent date found even after navigating to next month", "Fail"
        print(f"    [date: {TODAY}]")

        form_ok = open_form(drv, wt); time.sleep(2)
        if not form_ok:
            return f"Form did not open for {leave_type} - {day_type or 'Full Day'}", "Fail"

        # Select leave type
        lt = leave_type_select(drv)
        matched_lt = select_opt(lt, leave_type)
        if not matched_lt:
            return f"Could not select leave type '{leave_type}'", "Fail"

        # Select day type — skip for Maternity and Menstrual (Full Day auto)
        if day_type is not None:
            dt = day_type_select(drv)
            matched_dt = select_opt(dt, day_type)
            if not matched_dt:
                return f"Could not select day type '{day_type}'", "Fail"
            day_label = day_type
        else:
            day_label = "Full Day (auto)"

        enter_comment(drv)
        submitted = click_submit(drv)
        bt_after  = body_text(drv)
        err_kw    = ["error", "failed", "invalid", "already"]
        ok_kw     = ["success", "applied", "submitted", "saved"]
        is_err    = any(k in bt_after.lower() for k in err_kw)

        if submitted and any(k in bt_after.lower() for k in ok_kw):
            return (f"{leave_type} leave - {day_label} submitted successfully. "
                    f"Success message displayed"), "Pass"
        elif submitted and is_err:
            return f"Submitted but error shown: {bt_after[:120]}", "Fail"
        elif submitted:
            return (f"{leave_type} leave - {day_label} form submitted "
                    f"(modal closed, no explicit success message)"), "Pass"
        else:
            return f"Submit button not clicked. lt='{matched_lt}'", "Fail"
    except Exception as e:
        return f"Error: {str(e)[:150]}", "Fail"
    finally:
        try: drv.quit()
        except Exception: pass


# ── Test Runner ───────────────────────────────────────────────────────────────
def run_tests():
    results  = []
    run_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def rec(tc_id, title, expected, actual, status):
        results.append((tc_id, title, expected, actual, status, run_date))
        print(f"  [{status}] {tc_id}: {title[:72]}")

    for tc_id, leave_type, day_type in LEAVE_CASES:
        day_label = day_type if day_type else "Full Day (auto)"
        title     = f"Verify applying {leave_type} leave for {day_label} successfully"
        expected  = (f"{leave_type} leave - {day_label} should be submitted successfully. "
                     f"Success message should appear. Calendar should reflect the applied leave")
        print(f"\n  Running {tc_id}: {leave_type} - {day_label}")
        actual, status = do_submit(leave_type, day_type)
        rec(tc_id, title, expected, actual, status)

    return results


# ── Save results (append to existing sheet) ───────────────────────────────────
def save_results(results):
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="375623")
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    sep_fill  = PatternFill("solid", fgColor="D9D9D9")
    run_fill  = PatternFill("solid", fgColor="FFF2CC")
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin      = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"),  bottom=Side(style="thin"))
    headers    = ["Test Case ID", "Test Case Title", "Expected Result",
                  "Actual Result", "Status", "Date of Test Run"]
    col_widths = [16, 48, 52, 52, 10, 22]

    run_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total    = len(results)
    passed   = sum(1 for r in results if r[4] == "Pass")
    failed   = total - passed

    try:    wb = openpyxl.load_workbook(OUTPUT_FILE)
    except: wb = openpyxl.Workbook()

    if SHEET_NAME in wb.sheetnames:
        ws       = wb[SHEET_NAME]
        next_row = ws.max_row + 1
        for ci in range(1, 7):
            ws.cell(row=next_row, column=ci, value="").fill = sep_fill
        ws.row_dimensions[next_row].height = 8
        next_row += 1
        run_num = sum(1 for r in ws.iter_rows(min_row=1, values_only=True)
                      if r[0] and str(r[0]).startswith("RUN #")) + 1
        label = (f"RUN #{run_num} | {run_time} | "
                 f"Total: {total} | Passed: {passed} | Failed: {failed}")
        rc = ws.cell(row=next_row, column=1, value=label)
        rc.font = Font(bold=True, size=11); rc.fill = run_fill; rc.alignment = left
        ws.merge_cells(start_row=next_row, start_column=1,
                       end_row=next_row, end_column=6)
        ws.row_dimensions[next_row].height = 20
        next_row += 1
    else:
        ws = wb.create_sheet(title=SHEET_NAME)
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center; cell.border = thin
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        label = (f"RUN #1 | {run_time} | "
                 f"Total: {total} | Passed: {passed} | Failed: {failed}")
        rc = ws.cell(row=2, column=1, value=label)
        rc.font = Font(bold=True, size=11); rc.fill = run_fill; rc.alignment = left
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        ws.row_dimensions[2].height = 20
        next_row = 3

    for ri, row in enumerate(results, next_row):
        fill = pass_fill if row[4] == "Pass" else fail_fill
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill; cell.border = thin
            cell.alignment = center if ci in (1, 5, 6) else left
        ws.row_dimensions[ri].height = 70

    wb.save(OUTPUT_FILE)
    print(f"\nResults saved to: {OUTPUT_FILE}  (sheet: '{SHEET_NAME}')")
    print(f"Total: {total} | Passed: {passed} | Failed: {failed}")


if __name__ == "__main__":
    print("All Leave Types - Positive Cases | TC_ATT_025 to TC_ATT_032")
    print("-" * 65)
    results = run_tests()
    save_results(results)

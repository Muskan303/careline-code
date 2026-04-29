"""
Script: add_withdraw_leave_testcases.py
Purpose: Add a new tab 'Leave - Withdraw' with positive manual test cases
         for the Leave Withdrawal feature (via Tickets page) to
         Profile_BasicInfo_Manual_TestCases_v6.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE      = "Profile_BasicInfo_Manual_TestCases_v6.xlsx"
OUT_FILE  = "Profile_BasicInfo_Manual_TestCases_v6_updated.xlsx"
SHEET     = "Leave - Withdraw"

# ── Styles ────────────────────────────────────────────────────────────────────
hdr_font   = Font(bold=True, color="FFFFFF", size=11)
hdr_fill   = PatternFill("solid", fgColor="375623")   # dark green (matches existing sheets)
pass_fill  = PatternFill("solid", fgColor="C6EFCE")   # light green for positive cases
title_fill = PatternFill("solid", fgColor="D9E1F2")   # light blue for section title row
center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
left       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin       = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

HEADERS    = ["Test Case ID", "Test Case Title", "Pre-Conditions",
              "Test Steps", "Expected Result", "Status"]
COL_WIDTHS = [16, 45, 42, 60, 52, 12]

# ── Test Cases ────────────────────────────────────────────────────────────────
# Format: (tc_id, title, pre_conditions, steps, expected_result)
TEST_CASES = [
    (
        "TC_WD_001",
        "Verify ticket is created in Tickets page after applying a leave",
        "1. User is logged in.\n2. User has applied at least one leave (any leave type).\n3. Tickets page is accessible at:\nhttp://gcplcareline.girnarsoft.com/ticket/manage/listing?f=type.opentickets%2Borderby.lastactivitytime%2Bdirection.desc",
        "1. Apply a leave from the Attendance page.\n2. Navigate to the Tickets page URL.\n3. Observe the ticket listing.",
        "A new ticket corresponding to the applied leave should be visible in the Tickets page listing with correct details (leave type, date, employee name)."
    ),
    (
        "TC_WD_002",
        "Verify the 'Withdraw' button is visible on the leave ticket created by the logged-in user",
        "1. User is logged in.\n2. User has an open leave ticket in the Tickets page.\n3. Tickets page is open.",
        "1. Navigate to the Tickets page.\n2. Locate the ticket created for the user's own leave.\n3. Observe the available action buttons on the ticket.",
        "A 'Withdraw' button should be visible on the ticket that was created by the logged-in user."
    ),
    (
        "TC_WD_003",
        "Verify clicking the 'Withdraw' button opens a comment/confirmation dialog",
        "1. User is logged in.\n2. User has an open leave ticket visible in the Tickets page.\n3. 'Withdraw' button is visible on the ticket.",
        "1. Navigate to the Tickets page.\n2. Locate the user's own leave ticket.\n3. Click the 'Withdraw' button.",
        "A dialog/modal should open prompting the user to enter a comment before confirming the withdrawal."
    ),
    (
        "TC_WD_004",
        "Verify the comment field is present and accepts text input in the Withdraw dialog",
        "1. User is logged in.\n2. Withdraw dialog is open after clicking the 'Withdraw' button.",
        "1. Open the Withdraw dialog by clicking 'Withdraw' on a leave ticket.\n2. Click on the comment input field.\n3. Type a valid comment (e.g., 'Withdrawing leave due to change in plan').",
        "The comment field should be editable and accept the entered text without any errors."
    ),
    (
        "TC_WD_005",
        "Verify the 'Submit' button is present in the Withdraw dialog",
        "1. User is logged in.\n2. Withdraw dialog is open.",
        "1. Open the Withdraw dialog by clicking 'Withdraw' on a leave ticket.\n2. Observe the buttons available in the dialog.",
        "A 'Submit' button should be visible and enabled inside the Withdraw dialog."
    ),
    (
        "TC_WD_006",
        "Verify leave is successfully withdrawn after entering a comment and clicking Submit",
        "1. User is logged in.\n2. User has an open leave ticket in the Tickets page.\n3. Withdraw dialog is open.",
        "1. Navigate to the Tickets page.\n2. Click 'Withdraw' on the user's own leave ticket.\n3. Enter a valid comment (e.g., 'Withdrawing leave as plans changed').\n4. Click the 'Submit' button.",
        "The leave should be withdrawn successfully. A success message/confirmation should be displayed. The ticket status should update to reflect the withdrawal."
    ),
    (
        "TC_WD_007",
        "Verify the withdrawn leave ticket no longer appears in the Open Tickets list",
        "1. User is logged in.\n2. A leave has been successfully withdrawn (TC_WD_006 passed).",
        "1. After successfully withdrawing a leave, stay on or refresh the Tickets page.\n2. Check the Open Tickets listing.",
        "The withdrawn leave ticket should no longer appear in the Open Tickets list (it should be moved to closed/withdrawn status)."
    ),
    (
        "TC_WD_008",
        "Verify the Attendance calendar reflects the withdrawn leave (date reverts to Absent/Available)",
        "1. User is logged in.\n2. A leave has been successfully withdrawn.\n3. User navigates to the Attendance page.",
        "1. Withdraw a leave from the Tickets page.\n2. Navigate to the Attendance page.\n3. Locate the date for which the leave was withdrawn on the calendar.",
        "The date on the Attendance calendar should revert to its previous state (e.g., Absent) and should no longer show as a leave day."
    ),
    (
        "TC_WD_009",
        "Verify withdrawal works correctly for an Annual Leave ticket",
        "1. User is logged in.\n2. User has an open Annual Leave ticket in the Tickets page.",
        "1. Navigate to the Tickets page.\n2. Locate the Annual Leave ticket created by the user.\n3. Click 'Withdraw'.\n4. Enter a comment (e.g., 'Withdrawing Annual Leave').\n5. Click 'Submit'.",
        "Annual Leave should be withdrawn successfully. Success message should appear and the ticket should be removed from the Open Tickets list."
    ),
    (
        "TC_WD_010",
        "Verify withdrawal works correctly for a Bereavement Leave ticket",
        "1. User is logged in.\n2. User has an open Bereavement Leave ticket in the Tickets page.",
        "1. Navigate to the Tickets page.\n2. Locate the Bereavement Leave ticket created by the user.\n3. Click 'Withdraw'.\n4. Enter a comment (e.g., 'Withdrawing Bereavement Leave').\n5. Click 'Submit'.",
        "Bereavement Leave should be withdrawn successfully. Success message should appear and the ticket should be removed from the Open Tickets list."
    ),
    (
        "TC_WD_011",
        "Verify withdrawal works correctly for a Menstrual Leave ticket",
        "1. User is logged in.\n2. User has an open Menstrual Leave ticket in the Tickets page.",
        "1. Navigate to the Tickets page.\n2. Locate the Menstrual Leave ticket created by the user.\n3. Click 'Withdraw'.\n4. Enter a comment (e.g., 'Withdrawing Menstrual Leave').\n5. Click 'Submit'.",
        "Menstrual Leave should be withdrawn successfully. Success message should appear and the ticket should be removed from the Open Tickets list."
    ),
    (
        "TC_WD_012",
        "Verify withdrawal works correctly for a Maternity Leave ticket",
        "1. User is logged in.\n2. User has an open Maternity Leave ticket in the Tickets page.",
        "1. Navigate to the Tickets page.\n2. Locate the Maternity Leave ticket created by the user.\n3. Click 'Withdraw'.\n4. Enter a comment (e.g., 'Withdrawing Maternity Leave').\n5. Click 'Submit'.",
        "Maternity Leave should be withdrawn successfully. Success message should appear and the ticket should be removed from the Open Tickets list."
    ),
    (
        "TC_WD_013",
        "Verify withdrawal works correctly for an Election Leave ticket",
        "1. User is logged in.\n2. User has an open Election Leave ticket in the Tickets page.",
        "1. Navigate to the Tickets page.\n2. Locate the Election Leave ticket created by the user.\n3. Click 'Withdraw'.\n4. Enter a comment (e.g., 'Withdrawing Election Leave').\n5. Click 'Submit'.",
        "Election Leave should be withdrawn successfully. Success message should appear and the ticket should be removed from the Open Tickets list."
    ),
    (
        "TC_WD_014",
        "Verify withdrawal is successful when a long comment is entered",
        "1. User is logged in.\n2. User has an open leave ticket in the Tickets page.\n3. Withdraw dialog is open.",
        "1. Open the Withdraw dialog.\n2. Enter a long comment (e.g., 100+ characters: 'Withdrawing this leave because my plans have changed and I will be available on the requested date.').\n3. Click 'Submit'.",
        "The leave should be withdrawn successfully even with a long comment. Success message should be displayed."
    ),
    (
        "TC_WD_015",
        "Verify the Tickets page loads correctly and displays open leave tickets",
        "1. User is logged in.\n2. User has at least one open leave ticket.",
        "1. Navigate to:\nhttp://gcplcareline.girnarsoft.com/ticket/manage/listing?f=type.opentickets%2Borderby.lastactivitytime%2Bdirection.desc\n2. Observe the page content.",
        "The Tickets page should load successfully and display the list of open tickets sorted by last activity time in descending order."
    ),
]


# ── Build the sheet ───────────────────────────────────────────────────────────
def build_sheet():
    try:
        wb = openpyxl.load_workbook(FILE)
        print(f"Loaded existing file: {FILE}")
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        print(f"File not found — creating new workbook: {FILE}")

    # Remove sheet if it already exists (re-run safety)
    if SHEET in wb.sheetnames:
        del wb[SHEET]

    ws = wb.create_sheet(title=SHEET)

    # ── Header row ────────────────────────────────────────────────────────────
    for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── Section title row ─────────────────────────────────────────────────────
    section_cell = ws.cell(row=2, column=1,
        value="POSITIVE TEST CASES — Leave Withdrawal via Tickets Page")
    section_cell.font      = Font(bold=True, size=11, color="1F3864")
    section_cell.fill      = title_fill
    section_cell.alignment = left
    section_cell.border    = thin
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[2].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, (tc_id, title, pre, steps, expected) in enumerate(TEST_CASES, start=3):
        row_data = [tc_id, title, pre, steps, expected, ""]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = pass_fill
            cell.border    = thin
            cell.alignment = center if ci in (1, 6) else left
        ws.row_dimensions[ri].height = 90

    wb.save(OUT_FILE)
    print(f"\nDone! Saved as '{OUT_FILE}'")
    print(f"Sheet '{SHEET}' added with {len(TEST_CASES)} test cases.")
    print(f"\nNOTE: Close '{FILE}' in Excel, then rename '{OUT_FILE}' to '{FILE}' to replace it.")


if __name__ == "__main__":
    build_sheet()
